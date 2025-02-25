import os
import sys
import ctypes
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import platform
import shutil
import winreg
from tqdm import tqdm
import datetime
import psutil


def is_admin():
    try:
        return ctypes.windll.shell32.IsUserAnAdmin()
    except:
        return False


def convert_xls_to_xlsx(xls_path, excel=None):
    """使用Excel COM对象转换xls到xlsx，保留所有格式"""
    if platform.system() != 'Windows':
        raise RuntimeError("此功能需要Windows系统和Microsoft Excel。")
        
    abs_path = os.path.abspath(xls_path)
    temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "temp")
    
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)
    
    xlsx_path = os.path.join(temp_dir, f"{os.path.splitext(os.path.basename(xls_path))[0]}_temp.xlsx")
    
    should_quit = False
    if excel is None:
        try:
            import win32com.client
            excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = False
            excel.DisplayAlerts = False
            should_quit = True
        except ImportError:
            raise RuntimeError("请先安装 pywin32 库")
    
    try:
        wb = excel.Workbooks.Open(abs_path)
        wb.SaveAs(Filename=xlsx_path, FileFormat=51)
        wb.Close(SaveChanges=False)
        return xlsx_path
    finally:
        if should_quit and excel:
            try:
                excel.Application.Quit()
            except:
                pass


def get_processed_files():
    """获取今天已处理的文件列表"""
    log_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "processed_files.log")
    timestamp_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "last_process.timestamp")
    processed_files = set()
    
    # 检查上次处理的时间戳
    last_timestamp = 0
    if os.path.exists(timestamp_file):
        try:
            with open(timestamp_file, 'r') as f:
                last_timestamp = float(f.read().strip())
        except:
            last_timestamp = 0
    
    # 获取当前时间戳
    current_timestamp = datetime.datetime.now().timestamp()
    
    # 如果时间差小于24小时，则读取已处理文件列表
    if current_timestamp - last_timestamp < 24 * 3600:
        if os.path.exists(log_file):
            with open(log_file, 'r', encoding='utf-8') as f:
                processed_files = {line.strip() for line in f if line.strip()}
    else:
        # 如果超过24小时，清空记录
        if os.path.exists(log_file):
            os.remove(log_file)
    
    return processed_files


def log_processed_file(file_path):
    """记录处理过的文件"""
    log_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "processed_files.log")
    timestamp_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "last_process.timestamp")
    
    # 记录文件路径
    with open(log_file, 'a', encoding='utf-8') as f:
        f.write(f"{file_path}\n")
    
    # 更新时间戳
    with open(timestamp_file, 'w') as f:
        f.write(str(datetime.datetime.now().timestamp()))


def kill_excel_processes():
    """杀死所有Excel进程"""
    try:
        killed = False
        for proc in psutil.process_iter(['name']):
            try:
                if proc.name().lower() in ['excel.exe', 'microsoft excel']:
                    proc.kill()
                    killed = True
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass
        if killed:
            print("\nExcel进程已关闭")
    except:
        print("\n无法关闭Excel进程，请手动关闭Excel后重试")


def process_excel_files(input_number, path, output_dir):
    """处理Excel文件或文件夹"""
    # 先关闭所有Excel进程
    kill_excel_processes()
    
    # 获取今天已处理的文件
    processed_files = get_processed_files()
    
    # 预先创建Excel应用实例
    excel = None
    if platform.system() == 'Windows':
        try:
            import win32com.client
            excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = False
            excel.DisplayAlerts = False
        except:
            pass
    
    try:
        if os.path.isfile(path):
            abs_path = os.path.abspath(path)
            if path.lower().endswith(('.xls', '.xlsx')):
                if abs_path in processed_files:
                    print(f"跳过已处理文件: {path}")
                else:
                    with tqdm(total=1, desc="处理进度") as pbar:
                        process_single_excel(path, input_number, excel, output_dir)
                        log_processed_file(abs_path)
                        pbar.update(1)
                # 处理完成后打开输出文件夹
                open_folder(output_dir)
        elif os.path.isdir(path):
            # 获取所有Excel文件（使用绝对路径）
            all_files = []
            for file in os.listdir(path):
                if file.lower().endswith(('.xls', '.xlsx')):
                    full_path = os.path.abspath(os.path.join(path, file))
                    all_files.append(full_path)
            
            if not all_files:
                return

            # 分类文件
            to_process = []
            skipped = []
            for file_path in all_files:
                if file_path in processed_files:
                    skipped.append(os.path.basename(file_path))
                else:
                    to_process.append(file_path)

            # 显示处理信息
            total_files = len(all_files)
            print(f"\n找到 {total_files} 个Excel文件:")
            if skipped:
                print(f"- 已处理文件 ({len(skipped)}/{total_files}):")
                for file in skipped:
                    print(f"  跳过: {file}")
            
            if to_process:
                print(f"\n- 待处理文件 ({len(to_process)}/{total_files}):")
                with tqdm(total=len(to_process), desc="处理进度") as pbar:
                    for file_path in to_process:
                        print(f"\n正在处理: {os.path.basename(file_path)}")
                        process_single_excel(file_path, input_number, excel, output_dir)
                        log_processed_file(file_path)
                        pbar.update(1)
            
            # 处理完成后打开输出文件夹
            open_folder(output_dir)
    finally:
        # 清理Excel实例
        if excel:
            try:
                excel.Application.Quit()
            except:
                pass
        # 清理临时文件
        cleanup_temp_files()


def open_folder(path):
    """打开文件夹"""
    try:
        if platform.system() == 'Windows':
            os.startfile(path)
        elif platform.system() == 'Darwin':  # macOS
            os.system(f'open "{path}"')
        else:  # linux
            os.system(f'xdg-open "{path}"')
    except:
        pass


def process_single_excel(file_path, input_number, excel=None, output_dir=None):
    """处理单个Excel文件"""
    try:
        # 使用传入的输出目录
        if output_dir is None:
            if platform.system() == 'Windows':
                with winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                                  r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders") as key:
                    desktop_path = winreg.QueryValueEx(key, "Desktop")[0]
            else:
                desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            output_dir = os.path.join(desktop_path, "output")
        
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        file_name = os.path.basename(file_path)
        output_path = os.path.join(output_dir, file_name)

        if os.path.exists(output_path):
            os.remove(output_path)

        # 处理文件
        is_xls = file_path.lower().endswith('.xls')
        temp_xlsx_path = None

        try:
            if is_xls:
                if platform.system() != 'Windows':
                    return
                temp_xlsx_path = convert_xls_to_xlsx(file_path, excel)
                wb = load_workbook(temp_xlsx_path)
            else:
                wb = load_workbook(file_path)

            # 处理所有工作表
            sheets_processed = False
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if not any(cell.value is not None for row in ws.iter_rows() for cell in row):
                    continue

                empty_row = None
                for row in range(2, ws.max_row + 1):
                    if ws[f'Z{row}'].value is None:
                        empty_row = row
                        break

                if empty_row is None:
                    continue

                # 定义区域映射规则 - 每个区域对应2个数字
                rules = {
                    1: {
                        'area1': [6,1], 'area2': [5,6], 'area3': [4,5],
                        'area4': [3,4], 'area5': [2,3], 'area6': [1,2]
                    },
                    2: {
                        'area1': [1,2], 'area2': [6,1], 'area3': [5,6],
                        'area4': [4,5], 'area5': [3,4], 'area6': [2,3]
                    },
                    3: {
                        'area1': [2,3], 'area2': [1,2], 'area3': [6,1],
                        'area4': [5,6], 'area5': [4,5], 'area6': [3,4]
                    },
                    4: {
                        'area1': [3,4], 'area2': [2,3], 'area3': [1,2],
                        'area4': [6,1], 'area5': [5,6], 'area6': [4,5]
                    },
                    5: {
                        'area1': [4,5], 'area2': [3,4], 'area3': [2,3],
                        'area4': [1,2], 'area5': [6,1], 'area6': [5,6]
                    },
                    6: {
                        'area1': [5,6], 'area2': [4,5], 'area3': [3,4],
                        'area4': [2,3], 'area5': [1,2], 'area6': [6,1]
                    }
                }

                # 设置填充颜色 (RGB: 237, 112, 45)
                orange_fill = PatternFill(start_color='ED702D', end_color='ED702D', fill_type='solid')

                # 在找到的空行中填入数字
                ws[f'Z{empty_row}'].value = input_number

                print(f"\n处理输入数字: {input_number}")
                print("规则映射:")
                for i in range(1, 7):
                    print(f"区域{i}查找: {rules[input_number][f'area{i}']}")

                # 处理每个区域
                current_area = 1  # 当前处理的区域编号
                consecutive_empty = 0  # 连续空列计数
                last_non_empty = None  # 上一个非空列
                area_columns = {}  # 记录每个区域的列范围

                for col in range(2, ws.max_column + 1):  # 从B列开始
                    cell = ws.cell(row=empty_row, column=col)
                    
                    # 检查是否是空列
                    is_empty_col = True
                    for row in range(1, ws.max_row + 1):
                        if ws.cell(row=row, column=col).value is not None:
                            is_empty_col = False
                            break
                    
                    if is_empty_col:
                        consecutive_empty += 1
                        continue
                    
                    # 如果遇到非空列，且之前有连续的空列
                    if consecutive_empty > 0:
                        if last_non_empty is not None:  # 说明这是一个新区域的开始
                            # 记录上一个区域的列范围
                            area_columns[current_area] = (last_non_empty, col-consecutive_empty-1)
                            current_area += 1
                        consecutive_empty = 0
                    
                    last_non_empty = col
                    
                    # 处理当前单元格
                    if cell.value is not None:
                        cell_value = cell.value
                        if isinstance(cell_value, (int, float)):
                            cell_value = int(cell_value)
                            area_name = f'area{current_area}'
                            if current_area <= 6 and cell_value in rules[input_number][area_name]:
                                cell.fill = orange_fill
                                print(f"在区域{current_area}(列{col})找到匹配数字: {cell_value}")

                # 记录最后一个区域的列范围
                if last_non_empty is not None:
                    area_columns[current_area] = (last_non_empty, ws.max_column)

                print("\n区域列范围:")
                for area, (start, end) in area_columns.items():
                    col_letter_start = ws.cell(row=1, column=start).column_letter
                    col_letter_end = ws.cell(row=1, column=end).column_letter
                    print(f"区域{area}: {col_letter_start}-{col_letter_end}")

                sheets_processed = True

            if sheets_processed:
                if is_xls:
                    wb.save(temp_xlsx_path)
                    try:
                        save_as_xls(temp_xlsx_path, output_path, excel)
                    except Exception:
                        output_path = os.path.splitext(output_path)[0] + '.xlsx'
                        wb.save(output_path)
                else:
                    wb.save(output_path)

        finally:
            if temp_xlsx_path and os.path.exists(temp_xlsx_path):
                try:
                    os.remove(temp_xlsx_path)
                except:
                    pass

    except Exception:
        pass


def save_as_xls(xlsx_path, output_path, excel=None):
    """使用Excel COM对象将xlsx保存为xls"""
    should_quit = False
    if excel is None:
        try:
            import win32com.client
            excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = False
            excel.DisplayAlerts = False
            should_quit = True
        except ImportError:
            raise RuntimeError("请先安装 pywin32 库")
    
    try:
        wb = excel.Workbooks.Open(xlsx_path)
        wb.SaveAs(Filename=output_path, FileFormat=56)  # 56 is for .xls
        wb.Close(SaveChanges=False)
    finally:
        if should_quit and excel:
            try:
                excel.Application.Quit()
            except:
                pass


def cleanup_temp_files():
    """清理临时文件夹"""
    temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "temp")
    if os.path.exists(temp_dir):
        try:
            shutil.rmtree(temp_dir)
        except:
            pass


def check_and_install_dependencies():
    """检查并安装所需的库"""
    required_packages = {
        'openpyxl': 'openpyxl',
        'tqdm': 'tqdm',
        'pywin32': 'pywin32',
        'psutil': 'psutil'
    }
    
    import subprocess
    import sys
    
    def install_package(package_name):
        print(f"正在安装 {package_name}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package_name])
        print(f"{package_name} 安装完成")
    
    # 检查并安装缺失的包
    missing_packages = []
    for module, package in required_packages.items():
        try:
            __import__(module)
        except ImportError:
            missing_packages.append(package)
    
    if missing_packages:
        print("检测到缺少必要的库，正在自动安装...")
        try:
            for package in missing_packages:
                install_package(package)
            print("\n所有必要的库已安装完成")
        except Exception as e:
            print(f"\n安装库时出错: {str(e)}")
            print("请手动运行以下命令安装所需库:")
            for package in missing_packages:
                print(f"pip install {package}")
            input("\n按回车键退出...")
            sys.exit(1)


def select_path():
    """显示文件/文件夹选择对话框"""
    try:
        import tkinter as tk
        from tkinter import filedialog
        
        # 创建隐藏的主窗口
        root = tk.Tk()
        root.withdraw()
        
        # 显示选择对话框
        print("\n请选择Excel文件或文件夹...")
        path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xls;*.xlsx"), ("所有文件", "*.*")]
        )
        
        # 如果用户没有选择文件，提示选择文件夹
        if not path:
            print("\n未选择文件，请选择文件夹...")
            path = filedialog.askdirectory(title="选择文件夹")
        
        # 如果用户仍然没有选择，使用当前目录
        if not path:
            print("\n未选择路径，将处理当前目录")
            path = "."
            
        return path
        
    except Exception as e:
        print(f"\n打开文件选择对话框失败: {str(e)}")
        print("将使用命令行输入方式")
        path = input("\n请输入Excel文件或文件夹的路径 (直接回车处理当前目录): ").strip()
        if not path:
            path = "."
        return path


def select_output_path():
    """选择输出路径"""
    try:
        import tkinter as tk
        from tkinter import filedialog
        
        # 创建隐藏的主窗口
        root = tk.Tk()
        root.withdraw()
        
        # 获取默认桌面路径
        if platform.system() == 'Windows':
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                              r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders") as key:
                default_path = os.path.join(winreg.QueryValueEx(key, "Desktop")[0], "output")
        else:
            default_path = os.path.join(os.path.expanduser("~"), "Desktop", "output")
        
        # 显示选择对话框
        print("\n请选择输出文件夹 (取消则使用桌面默认路径)...")
        path = filedialog.askdirectory(
            title="选择输出文件夹",
            initialdir=os.path.dirname(default_path)
        )
        
        # 如果用户取消选择，使用默认路径
        if not path:
            path = default_path
            print(f"\n使用默认输出路径: {path}")
        
        # 确保输出目录存在
        if not os.path.exists(path):
            os.makedirs(path)
            
        return path
        
    except Exception as e:
        print(f"\n选择输出路径失败: {str(e)}")
        # 使用默认桌面路径
        default_path = os.path.join(os.path.expanduser("~"), "Desktop", "output")
        if not os.path.exists(default_path):
            os.makedirs(default_path)
        print(f"\n使用默认输出路径: {default_path}")
        return default_path


if __name__ == "__main__":
    try:
        # 检查并安装依赖
        check_and_install_dependencies()
        
        if not is_admin():
            ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, " ".join(sys.argv), None, 1)
        else:
            # 选择输出路径
            output_path = select_output_path()
            
            while True:
                try:
                    input_number = int(input("请输入一个数字(1-6): "))
                    if 1 <= input_number <= 6:
                        break
                    else:
                        print("请输入1到6之间的数字")
                except ValueError:
                    print("请输入有效的数字")

            # 使用文件选择对话框
            path = select_path()

            if os.path.exists(path):
                # 修改 process_excel_files 调用，传入输出路径
                process_excel_files(input_number, path, output_path)
            else:
                print(f"找不到路径: {path}")
    except Exception as e:
        print(f"\n程序出错: {str(e)}")
    finally:
        # 等待用户确认后退出
        input("\n按回车键退出...")